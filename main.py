#pip install python-dateutil
from datetime import datetime,timedelta, date
from dateutil import parser
import os
import openpyxl
import logging
import configparser
import shutil
import re
import openpyxl.utils
import openpyxl.utils.exceptions

def cargar_configuracion():
        global ruta_carpeta, archivo_excel_year, archivo_excel_base
        global variables, variables_ruta, config_ruta, RUTAS, POSICION_COLUMNA, config, ruta_actual, NOMBRES

        config = configparser.ConfigParser()
        ruta_actual = os.path.dirname(os.path.abspath(__file__))

        variables_ruta = os.path.join(ruta_actual, 'variables.ini')
        config_ruta = os.path.join(ruta_actual, 'config.ini')
        
        ruta_carpeta = r'/mnt/ZAM-ID-APP2/i'
        carpeta_excel = os.path.join(ruta_actual, "Bitacora_APP")

        try:
            config.read(config_ruta)
            RUTAS = config.get('CARPETAS', 'rutas').split(',')
            POSICION_COLUMNA = config.get('CARPETAS', 'posicionColumna').split(',')
            NOMBRES = config.get('CARPETAS', 'nombres').split(',')
        except (configparser.Error, KeyError) as e:
            print(f"Error al cargar la configuracion: {e}")

        archivo_excel_base = os.path.join(ruta_actual, "Bitacora_APP.xlsx")
        archivo_excel_year = os.path.join(carpeta_excel, f"Bitacora_APP_{now.year}.xlsx")

        variables = configparser.ConfigParser()
        try:
            variables.read(variables_ruta)
        except configparser.Error as e:
            print(f"Error al leer el archivo de variables: {e}")

def marcar(r,subcarpeta_,archivo1,archivo2,posicion_,incremento):
    global marcado
    if incremento == posicion_: 
        subRuta1 = os.path.join(ruta_carpeta, r, subcarpeta_, archivo1)
        
        if os.path.exists(subRuta1):
            marcado[posicion_] += 1

        if archivo2 != "":
            subRuta2 = os.path.join(ruta_carpeta, r, subcarpeta_, archivo2)
            if os.path.exists(subRuta2):
                marcado[posicion_] += 1

def verificar_ruta(ruta):
        if not os.path.exists(ruta):
            raise FileNotFoundError(f"La ruta: {ruta}, no existe")
        if not os.access(ruta, os.R_OK):
            raise PermissionError(f"No tiene permisos de lectura en la ruta: {ruta}")
        if not os.access(ruta, os.W_OK):
            raise PermissionError(f"No tienes permisos de escritura en la ruta: {ruta}")

def generacionCarpeta():
        try: 
            verificar_ruta(archivo_excel_base)

            if not os.path.exists(archivo_excel_year):
                shutil.copy2(archivo_excel_base, archivo_excel_year)
                print(f"Se ha generado bitacora para el año {fechaFija.year}")
                
        except FileNotFoundError as e:
            print(f"Error al generar la carpeta de la bitácora: Archivo no encontrado - {e}")

        except PermissionError as e:
            print(f"Error al generar la carpeta de la bitácora: Permiso denegado - {e}")

        except OSError as e:
            print(f"Error al generar la carpeta de la bitácora: Error del sistema - {e}")

        except Exception as e:
            print(f"Error inesperado al generar la carpeta de la bitácora: {e}")

def convFecha(f):
    f_menos_uno = f + timedelta(days=-1)  

    mes = str(f.month).zfill(2)
    dia = str(f.day).zfill(2)

    mes_menos_uno = str(f_menos_uno.month).zfill(2)
    dia_menos_uno = str(f_menos_uno.day).zfill(2)
   
    return mes,dia,dia_menos_uno,mes_menos_uno

def calcular_posicion_columna(fecha):
    FECHA_REFERENCIA = datetime(2024, 1, 5).date() 

    if isinstance(fecha, datetime):
        fecha = fecha.date() 

    diferencia_semana = (fecha - FECHA_REFERENCIA).days // 7

    return 3 + diferencia_semana

def Bitacora_APP():
    global RUTAS, fechaFija, marcado, NOMBRES, ruta_actual
    global day, month, year

    ruta_logging = os.path.join(ruta_actual, "archivo_log.txt")

    logging.basicConfig(filename=ruta_logging, level=logging.ERROR)

    generacionCarpeta()

    fechaFija = datetime.strptime(f"{day}/{month}/{year}", '%d/%m/%Y').date()

    copiarColumna = True

    mes = None
    dia = None
    dia_menos_uno = None
    mes_menos_uno = None

    marcado = [0] * (len(RUTAS) + 1)

    for i,rutas in enumerate(RUTAS):
        for n in range(2,9):
            mes, dia,dia_menos_uno,mes_menos_uno = convFecha(fechaFija + timedelta(days=-(7-n)))
            #CDSF
            marcar( rutas,
                f'{fechaFija.year}_{mes}_{dia}',
                    f"Lobo_RH_APP_ZAM-ID-PDC_{dia_menos_uno}-{mes}-{fechaFija.year}.7z",
                    f"Lobo_RH_APP_ZAM_SV_CDSF_{dia}-{mes}-{fechaFija.year}.7z",
                    0,i)
            #BUGZILLA
            marcar( rutas,
                    f"",
                    f"bugzilla-{fechaFija.year}{mes}{dia}-03-00.tar.gz",
                    f"bugs-{fechaFija.year}{mes}{dia}-03-00.tar.gz",
                    1,i)
            #GIT
            marcar( rutas,
                    f"",
                    f"{fechaFija.year}{mes}{dia}-01-00-dump_gitlab_backup.tar",
                    f"",
                    2,i)
            #Proyectos_SVN
            marcar( rutas,
                    f"",
                    f"Proyectos_svn_full_backup-{fechaFija.year}{mes}{dia}.rar",
                    f"",
                    3,i)
            #Sial_SVN
            marcar( rutas,
                    f"",
                    f"SIAL_svn_full_backup-{fechaFija.year}{mes}{dia}.rar",
                    f"",
                    4,i)
            #Sial_PDC
            marcar( rutas,
                    f"{fechaFija.year}_{mes}_{dia}",
                    f"Lobo_RH_APP_ZAM-ID-PDC_{dia_menos_uno}-{mes_menos_uno}-{fechaFija.year}.7z",
                    f"",
                    5,i)
            #WWW LOBO
            marcar( rutas,
                    f"",
                    f"lobos-{fechaFija.year}{mes}{dia}-23-01.tar.gz",
                    f"",
                    6,i)
            #WWW PDV
            marcar( rutas,
                    f"",
                    f"pdv-{fechaFija.year}{mes}{dia}-23-01.tar.gz",
                    f"",
                    7,i)
            #WWW ICCAS
            marcar( rutas,
                    f"",
                    f"iccas-{fechaFija.year}{mes}{dia}-23-01.tar.gz",
                    f"",
                    8,i)
            #WWW ZAMGOB
            marcar( rutas,
                    f"",
                    f"DATA/zamgob-{fechaFija.year}{mes}{dia}-23-01.tar.gz",
                    f"DB/zamoragob-{fechaFija.year}{mes}{dia}-23-01.tar.gz",
                    9,i)
            #COMPARTIDA
            marcar( rutas,
                    f"",
                    f"RESPALDO-compartida_zam-id-gab_{dia}-{mes}-{fechaFija.year}.rar",
                    f"",
                    10,i)        
    try:
        verificar_ruta(archivo_excel_year)
        workbook = openpyxl.load_workbook(archivo_excel_year)
        sheet = workbook.active

        columna_calculada = calcular_posicion_columna(fechaFija)

        if (copiarColumna):
            copiar_fila(7, columna_calculada - 1, 7, columna_calculada, sheet)
            fecha_formateada = fechaFija.strftime('%d/%m/%Y') 

            sheet.cell(row=7, column=columna_calculada).value = fecha_formateada
            sheet.cell(row=7, column=columna_calculada).number_format = 'DD/MM/YYYY'  
            copiarColumna = False

        for i,valor in enumerate(POSICION_COLUMNA):
            if (marcado[i] > 0):
                copiar_fila(100,1,int(valor),calcular_posicion_columna(fechaFija), sheet)
                sheet.cell(row=int(valor), column=calcular_posicion_columna(fechaFija)).value = "1";marcado[i] = 0

        if (not pruebas):
            with open(variables_ruta, 'w') as archivo_config:
                variables.write(archivo_config)

        if (not pruebas):
            workbook.save(archivo_excel_year)
        workbook.close()
        
    except FileNotFoundError as e: 
        print(f"No se pudo abrir el archivo de excel: {e}")
    except openpyxl.utils.exceptions.InvalidFileException as e:
        print(f"El archivo excel es invalido: {e}")
    except Exception as e:
        print(f'No se pudo abrir excel {e}')

def copiar_fila(origen_row,origen_col,detino_row,detino_col, sh):
    cell_1 = sh.cell(row=origen_row, column=origen_col)
    cell_2 = sh.cell(row=detino_row, column=detino_col)
    cell_2._style = cell_1._style 

first_date = datetime(2024, 1, 5).date()

day = first_date.day
month = first_date.month
year = first_date.year

now = datetime.now()

while (first_date <= now.date()):
    pruebas = False

    cargar_configuracion()
    Bitacora_APP()
    
    first_date += timedelta(days=7)

    day = first_date.day
    month = first_date.month
    year = first_date.year

##################################################PERFILES#######################################################
def detectar_formato_fecha(fecha):
    try:
        fecha_objeto = parser.parse(fecha)
        return fecha_objeto.strftime("%d/%m/%Y")
    except ValueError as e:
        raise ValueError(f"Formato de fecha no reconocido: {fecha}. Error: {str(e)}")

def leer_ultimas_lineas(nombre_archivo, num_lineas):
    global fecha_log_lista, mes_log, dia_log, anio_log

    mes_log = None
    dia_log = None
    anio_log = None

    try:
        with open(nombre_archivo, 'r') as archivo:
            lineas = archivo.readlines()
            ultimas_lineas = lineas[-num_lineas:] 
            
            leer_fecha_log = lineas[5:6]
            leer_fecha_log = leer_fecha_log[0][12:]

            leer_fecha_log = leer_fecha_log.strip() 
            leer_fecha_log = re.sub(r'[^\w\s:]', '', leer_fecha_log) 
            
            mes_log, dia_log,anio_log = convertir_formato(leer_fecha_log)

            fecha_convertida = detectar_formato_fecha(leer_fecha_log)
            fecha_log_lista = fecha_convertida.split('/')

            contenido = ''.join(ultimas_lineas)
            
            return contenido
        
    except FileNotFoundError as e:
        print(f"No se pudo abrir el archivo de log: {e}")
    except PermissionError as e:
        print(f"No tienes permisos para leer el archivo de log: {e}")
    except IndexError as e:
        print(f"Índice fuera de rango al intentar leer la fecha del archivo de log: {e}")
    except Exception as e:
        print(f"Ocurrió un error inesperado al leer el archivo de log: {e}")
    
def limpiar_cadena(fecha_str):
    return re.sub(r'[^\w\s:]', '', fecha_str.strip())

def convertir_formato(fecha_str):
    fecha_limpiada = limpiar_cadena(fecha_str)
    
    fecha_obj = datetime.strptime(fecha_limpiada, '%a %b %d %H:%M:%S %Y')
    
    mes = fecha_obj.month
    dia = fecha_obj.day
    anio = fecha_obj.year
    
    return mes, dia, anio

def marcar_perfiles():
    LOG_PERFILES = ('robocopy-ABC.log','robocopy-LOBO.log','robocopy-NAS-LOBO.log')
    LOG_N = (10,10,10)
    LOG_V = ('Dir','Files')#,'Bytes')
    
    marcarPerfil = 0

    print("\nPERFILES\n")

    for i, log_file in enumerate(LOG_PERFILES):
        for j, valor in enumerate(LOG_V):
            try:
                perfiles = os.path.join(ruta_carpeta, log_file)
                
                contenido = leer_ultimas_lineas(perfiles, LOG_N[i] - j)

                fechaLog = datetime(anio_log, mes_log, dia_log).date() 

                if (fechaLog.weekday() != 4):
                    raise ValueError("La fecha no es viernes...")
                            
                contenido_FAILED = contenido[51:60]
                print(f"Errores encontrados {LOG_V[j]}: {int(contenido_FAILED)}")
                
                if (int(contenido_FAILED) == 0):
                    marcarPerfil += 1
            except FileNotFoundError as e:
                print(f"No se pudo abrir el archivo de log {log_file}: {e}")
            except PermissionError as e:
                print(f"No tienes permisos para leer el archivo de log {log_file}: {e}")
            except ValueError as e:
                print(f"Error en la fecha en el archivo de log {log_file}: {e}")
            except Exception as e:
                print(f"Error inesperado al procesar el archivo de log {log_file}: {e}")
    try:
        for m,valor in enumerate(RUTAS):
            if valor == "PERFILES":
                marcar_perfiles_n = int(POSICION_COLUMNA[m])
                break

        workbook = openpyxl.load_workbook(archivo_excel_year)
        sheet = workbook.active

        if marcarPerfil == 6:
            copiar_fila(100,1,marcar_perfiles_n,calcular_posicion_columna(fechaLog), sheet)
            sheet.cell(row=marcar_perfiles_n, column=calcular_posicion_columna(fechaLog)).value = "1"
            print("Marcando PERFILES = 1");marcarPerfil = 0

        elif marcarPerfil > 0 and marcarPerfil < 6:
            copiar_fila(100,1,marcar_perfiles_n,calcular_posicion_columna(fechaLog), sheet)
            sheet.cell(row=marcar_perfiles_n, column=calcular_posicion_columna(fechaLog)).value = "-"
            print("Marcando PERFILES = - \nHubo algunos errores.");marcarPerfil = 0

        else:
            print(f"No se marco, revisar errores en log")

        print("Guardando...")

        if (not pruebas):
            workbook.save(archivo_excel_year)
        workbook.close()
        
    except FileNotFoundError as e:
        print(f"No se pudo abrir el archivo de excel {archivo_excel_year}: {e}")
    except openpyxl.utils.exceptions.InvalidFileException as e:
        print(f"El archivo excel es inválido: {e}")
    except Exception as e:
        print(f"Ocurrió un error inesperado al procesar la bitácora: {e}")
    print("Termino de generacion de Bitacora_APP.")

print("Iniciando Generacion de Bitacora_APP")

marcar_perfiles()